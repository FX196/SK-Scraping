import facebook
from facebook import GraphAPIError
from openpyxl import Workbook
from openpyxl.utils.exceptions import IllegalCharacterError
from datetime import datetime, timezone
import urllib.request
import time
import inspect
import os

filename = inspect.getframeinfo(inspect.currentframe()).filename
file_path = os.path.dirname(os.path.abspath(filename))
token = "EAACXp46VJZBgBAEVQDdhn4ZCFyhkHMF2bV9Bz9U4YwmaoD4Rhv1Qgd1lXHeMZCxkW0TT8XEpQ2ZAI8dGI5dfreZBhoAnk9BRFnUF7GaHLYSUJA8FKBobS6UOeBD76UNwINqHdLKI2ghJYl98KTJtZB0GzQw9cGk7UZD"
post_id = "1049379735171791_1578479555595137"


class FacebookScraper:
    def __init__(self, token, post_id):
        self.graph = facebook.GraphAPI(access_token=token, version=3.0)
        self.post_id = post_id
        self.deadline = datetime(2018, 8, 13, 15, 59, 59, tzinfo=timezone.utc)
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Facebook"
        self.next_row = 1
        self.row_dict = {1: 1}
        print("Connected to Facebook")
        self.start_time = time.time()

    def get_comments(self):
        print("Grabbing comments")
        comments = self.graph.get_connections(self.post_id, 'comments', limit=2000)
        print("Grabbed %d comments" % len(comments['data']))
        return comments  # returns dict of comments

    def parse_comment(self, comment_id):
        comment = self.graph.get_object(comment_id, fields='message,attachment,created_time')
        # comment_id = comment_id
        time.sleep(10)
        comment_text = comment['message']
        if 'attachment' in comment.keys():
            comment_img_src = comment['attachment']['media']['image']['src']
        else:
            comment_img_src = None
        comment_time = comment['created_time']

        return comment_id, comment_text, comment_time, comment_img_src

    def insert_to_worksheet(self):
        directory = os.path.dirname(file_path + "/Facebook_Pictures")
        if not os.path.exists(directory):
            os.makedirs(directory)
        comments = self.get_comments()
        total_length = len(comments['data'])
        i = total_length-1
        tries = 0
        while i > 0:
            elapsed_time = time.time() - self.start_time
            try:
                parsed = self.parse_comment(comments['data'][i]['id'])
                self.save()
            except GraphAPIError as e:
                print("\n", str(e), "Elapsed Time: %02d:%02d:%02d" % (
                    elapsed_time // 3600, (elapsed_time % 3600) // 60, elapsed_time % 60))
                if tries < 5:
                    time.sleep(180)
                    tries += 1
                else:
                    time.sleep(1200)
                    tries = 0
                continue
            tries = 0
            if len(parsed[1]) < 30 or datetime.strptime(parsed[2], "%Y-%m-%dT%H:%M:%S%z") > self.deadline:
                continue
            text_data, img_data = list(parsed[:-1]), parsed[-1]
            for j in range(3):
                try:
                    self.ws.cell(row=self.next_row, column=(j + 1), value=text_data[j])
                except IllegalCharacterError:
                    pass
            if img_data:
                # save image
                img_url = img_data
                urllib.request.urlretrieve(img_url,
                                           file_path + "/Facebook_Pictures/%s.jpg" % (
                                               parsed[0]))
                # insert to worksheet
                self.ws.cell(row=self.next_row, column=4, value=img_url)
            print("\r", "%d out of %d comments processed" % (total_length-i, total_length),
                  "Elapsed Time: %02d:%02d:%02d" % (
                      elapsed_time // 3600, (elapsed_time % 3600) // 60, elapsed_time % 60),
                  end="")
            self.next_row += 1
            if self.next_row % 100 == 0:
                print("\nProcessed %d rows" % self.next_row)
                self.save(self.next_row)
            i -= 1

    def save(self, checkpoint=None):
        if not checkpoint:
            self.wb.save('Facebook.xlsx')
        else:
            self.wb.save('Facebook-checkpoint-%d.xlsx' % checkpoint)


if __name__ == "__main__":
    FBScraper = FacebookScraper(token, post_id)
    FBScraper.insert_to_worksheet()
    FBScraper.save()
